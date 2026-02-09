#!/usr/bin/env python3
"""
Assign each lead to its *nearest* dealer (and optionally only keep leads within a radius),
then export an Excel report.

Why this is fast:
- Geocodes each unique (city, province/state, country) at most once using geo_city_cache.json
- Computes nearest dealer via pure distance math (no extra geocoding inside the dealer loop)

Inputs:
- leads_export.json from your existing exporter/GUI
- Dealers from odoo_connector.DEALER_LOCATIONS (must have Latitude/Longitude)

Outputs:
- Excel workbook with:
  - "Summary": counts of assigned leads per dealer (within radius filter)
  - "Assigned": one row per lead with nearest dealer + distance_km

Notes:
- If a lead has no city/prov (or geocoding fails), it's skipped (reported in console).
- This assigns *one* dealer per lead (the closest). It does NOT list "all dealers within 50 km".

To run this 
python3 lead_nearest_dealer_report_v3.py \
  --leads leads_export.json \
  --use-driving \
  --max-hours 2 \
  --fetch-odoo \
  --out lead_nearest_2h.xlsx


"""

from __future__ import annotations

import argparse
import json
import math
import time
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook

from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

from odoo_connector import DEALER_LOCATIONS, haversine_distance, connect_odoo, ODOO_DB, ODOO_PASSWORD

GEO_CACHE_PATH = Path("geo_city_cache.json")
_geolocator = Nominatim(user_agent="WavcorLeadNearestDealer/1.0")

DEFAULT_RADIUS_KM = 50.0
DEFAULT_TOPK = 5
OSRM_BASE_URL = "https://router.project-osrm.org"

LEAD_COLUMNS = [
    "id",
    "name",
    "type",
    "stage_name",
    "contact_name",
    "partner_name",
    "email_from",
    "phone",
    "mobile",
    "city",
    "province_state_code",
    "province_state_name",
    "province",
    "zip",
    "country_name",
    "country",
    "create_date",
    "write_date",
    "user_name",
    "team_name",
]

def get_lead_id(lead: Dict[str, Any]) -> Optional[int]:
    """Extract crm.lead ID from an exported lead dict (handles different key casings)."""
    candidates = ["id", "lead_id", "crm_id", "odoo_id"]
    for cand in candidates:
        if cand in lead and lead[cand]:
            try:
                return int(lead[cand])
            except Exception:
                pass
        for k, v in lead.items():
            if isinstance(k, str) and k.lower() == cand and v:
                try:
                    return int(v)
                except Exception:
                    pass
    return None


def fetch_lead_fields_from_odoo(lead_ids: List[int]) -> Tuple[Dict[int, str], Dict[int, List[int]]]:
    """Fetch email_from and tag_ids for the given crm.lead IDs via XML-RPC."""
    uid, models = connect_odoo()
    if not uid or not models:
        raise RuntimeError("Could not authenticate/connect to Odoo (connect_odoo failed).")

    email_by_id: Dict[int, str] = {}
    tagids_by_id: Dict[int, List[int]] = {}

    fields = ["id", "email_from", "tag_ids"]
    CHUNK = 200
    for i in range(0, len(lead_ids), CHUNK):
        chunk = lead_ids[i:i+CHUNK]
        recs = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            "crm.lead", "read",
            [chunk],
            {"fields": fields},
        )
        for r in recs or []:
            rid = int(r.get("id"))
            email_by_id[rid] = (r.get("email_from") or "").strip()
            tagids = r.get("tag_ids") or []
            # tag_ids is a list of ints
            tagids_by_id[rid] = [int(t) for t in tagids if t]
    return email_by_id, tagids_by_id


def fetch_tag_names_from_odoo(tag_ids: List[int]) -> Dict[int, str]:
    """Fetch crm.tag names for tag IDs."""
    if not tag_ids:
        return {}
    uid, models = connect_odoo()
    if not uid or not models:
        raise RuntimeError("Could not authenticate/connect to Odoo (connect_odoo failed).")

    name_by_id: Dict[int, str] = {}
    CHUNK = 400
    for i in range(0, len(tag_ids), CHUNK):
        chunk = tag_ids[i:i+CHUNK]
        recs = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            "crm.tag", "read",
            [chunk],
            {"fields": ["id", "name"]},
        )
        for r in recs or []:
            rid = int(r.get("id"))
            name_by_id[rid] = (r.get("name") or "").strip()
    return name_by_id

def load_leads_export(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("records"), list):
        return data["records"]
    if isinstance(data, list):
        return data
    raise ValueError(f"Unrecognized leads JSON format: {path}")


def _geo_key(city: str, prov: str, country: str) -> str:
    return f"{city.strip().lower()}|{prov.strip().lower()}|{country.strip().lower()}"


def _load_geo_cache() -> Dict[str, Tuple[float, float]]:
    if not GEO_CACHE_PATH.exists():
        return {}
    try:
        raw = json.loads(GEO_CACHE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    clean: Dict[str, Tuple[float, float]] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            try:
                if isinstance(v, (list, tuple)) and len(v) == 2 and v[0] is not None and v[1] is not None:
                    clean[k] = (float(v[0]), float(v[1]))
            except Exception:
                continue
    return clean


def _save_geo_cache(cache: Dict[str, Tuple[float, float]]) -> None:
    GEO_CACHE_PATH.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")

def _routes_cache_path() -> Path:
    return Path("route_duration_cache.json")


def _load_routes_cache() -> Dict[str, Dict[str, float]]:
    path = _routes_cache_path()
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if isinstance(raw, dict):
        return raw
    return {}


def _save_routes_cache(cache: Dict[str, Dict[str, float]]) -> None:
    _routes_cache_path().write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")


def _route_key(lat1: float, lon1: float, lat2: float, lon2: float) -> str:
    return f"{lat1:.6f},{lon1:.6f}->{lat2:.6f},{lon2:.6f}"


def _osrm_route_duration_s(
    lat1: float, lon1: float, lat2: float, lon2: float,
    cache: Dict[str, Dict[str, float]],
) -> Optional[float]:
    key = _route_key(lat1, lon1, lat2, lon2)
    if key in cache:
        return cache[key].get("duration_s")

    coords = f"{lon1},{lat1};{lon2},{lat2}"
    url = f"{OSRM_BASE_URL}/route/v1/driving/{coords}"
    params = urllib.parse.urlencode({"overview": "false", "alternatives": "false"})
    try:
        with urllib.request.urlopen(f"{url}?{params}", timeout=20) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        routes = payload.get("routes") or []
        if not routes:
            return None
        duration_s = float(routes[0].get("duration"))
        cache[key] = {"duration_s": duration_s}
        return duration_s
    except Exception:
        return None


def geocode_city_prov(
    city: str,
    prov: str,
    country: str,
    cache: Dict[str, Tuple[float, float]],
    attempts: int = 3,
    sleep_s: float = 1.05,
) -> Optional[Tuple[float, float]]:
    """
    Geocode city/prov/country, caching results. Returns (lat, lon) or None.
    Polite throttling for Nominatim via sleep_s between successful calls.
    """
    if not city or not prov:
        return None

    key = _geo_key(city, prov, country)
    if key in cache:
        return cache[key]

    query = f"{city}, {prov}, {country}"
    for i in range(1, attempts + 1):
        try:
            loc = _geolocator.geocode(query, timeout=10)
            if loc:
                cache[key] = (loc.latitude, loc.longitude)
                # polite throttle (only when we actually called out)
                time.sleep(sleep_s)
                return cache[key]
            return None
        except (GeocoderTimedOut, GeocoderServiceError):
            time.sleep(1.5 * i)
        except Exception:
            return None
    return None


def _dealer_name(d: Dict[str, Any]) -> str:
    return str(d.get("Location", "")).strip()


def _prepare_dealers() -> List[Tuple[str, float, float]]:
    dealers: List[Tuple[str, float, float]] = []
    for d in DEALER_LOCATIONS:
        name = _dealer_name(d)
        if not name:
            continue
        try:
            lat = float(d["Latitude"])
            lon = float(d["Longitude"])
        except Exception:
            continue
        dealers.append((name, lat, lon))
    if not dealers:
        raise SystemExit("No valid dealers with Latitude/Longitude found in DEALER_LOCATIONS.")
    return dealers


def _lead_city_prov_country(lead: Dict[str, Any], default_country: str = "Canada") -> Tuple[str, str, str]:
    city = (lead.get("city") or "").strip()
    prov = (
        (lead.get("province_state_code") or "")
        or (lead.get("province_state_name") or "")
        or (lead.get("province") or "")
    ).strip()
    country = (lead.get("country_name") or lead.get("country") or default_country).strip() or default_country
    return city, prov, country


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--leads", required=True, help="Path to leads_export.json")
    ap.add_argument("--radius-km", type=float, default=DEFAULT_RADIUS_KM, help="Only keep leads within this distance of their nearest dealer")
    ap.add_argument("--use-driving", action="store_true", help="Use OSRM driving time to choose nearest dealer")
    ap.add_argument("--max-hours", type=float, default=None, help="Only keep leads within this many driving hours of their nearest dealer (requires --use-driving)")
    ap.add_argument("--topk", type=int, default=DEFAULT_TOPK, help="When using --use-driving, only route the top K closest by straight-line distance")
    ap.add_argument("--out", default="lead_nearest_dealer.xlsx", help="Output .xlsx filename")
    ap.add_argument("--default-country", default="Canada", help="Fallback country name (default: Canada)")
    ap.add_argument("--fetch-odoo", action="store_true", help="Fetch email_from and tags from Odoo by lead ID (requires Odoo credentials in odoo_connector.py)")
    args = ap.parse_args()

    leads_path = Path(args.leads).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()

    leads = load_leads_export(leads_path)
    dealers = _prepare_dealers()

    # 1) Resolve coordinates for leads (cached by unique city/prov/country)
    cache = _load_geo_cache()
    cache_dirty = False

    resolved: List[Tuple[Dict[str, Any], float, float]] = []
    skipped_missing = 0
    skipped_geocode = 0
    unassigned_rows: List[Dict[str, Any]] = []

    # Build a set of unique keys first to minimize geocode calls
    unique_keys: Dict[str, Tuple[str, str, str]] = {}
    for lead in leads:
        city, prov, country = _lead_city_prov_country(lead, default_country=args.default_country)
        if not city or not prov:
            out = dict(lead)
            out["unassigned_reason"] = "missing_city_or_province"
            unassigned_rows.append(out)
            continue
        k = _geo_key(city, prov, country)
        if k not in unique_keys:
            unique_keys[k] = (city, prov, country)

    # Geocode only keys not already in cache
    keys_to_geocode = [(k, *unique_keys[k]) for k in unique_keys.keys() if k not in cache]
    if keys_to_geocode:
        print(f"Geocoding {len(keys_to_geocode)} unique city/province combinations (cached in {GEO_CACHE_PATH})...")
    for idx, (k, city, prov, country) in enumerate(keys_to_geocode, start=1):
        coords = geocode_city_prov(city, prov, country, cache)
        if coords:
            cache_dirty = True
        if idx % 25 == 0:
            print(f"  Geocoded {idx}/{len(keys_to_geocode)}...")

    if cache_dirty:
        _save_geo_cache(cache)

    # Now resolve each lead quickly from cache
    for lead in leads:
        city, prov, country = _lead_city_prov_country(lead, default_country=args.default_country)
        if not city or not prov:
            skipped_missing += 1
            continue
        k = _geo_key(city, prov, country)
        coords = cache.get(k)
        if not coords:
            out = dict(lead)
            out["unassigned_reason"] = "geocode_failed"
            unassigned_rows.append(out)
            skipped_geocode += 1
            continue
        lid = get_lead_id(lead)
        if lid is not None and not lead.get('id'):
            lead['id'] = lid
        resolved.append((lead, coords[0], coords[1]))

    print(f"Resolved coords for {len(resolved)}/{len(leads)} leads. Skipped missing city/prov: {skipped_missing}. Skipped geocode: {skipped_geocode}.")

    # 2) For each lead, find nearest dealer
    assigned_rows: List[Dict[str, Any]] = []
    counts: Dict[str, int] = {}

    radius_km = float(args.radius_km)
    use_driving = bool(args.use_driving)
    max_hours = args.max_hours
    topk = max(1, int(args.topk))
    routes_cache = _load_routes_cache() if use_driving else {}
    routes_cache_dirty = False

    for i, (lead, lat, lon) in enumerate(resolved, start=1):
        if use_driving:
            ranked = []
            for name, dlat, dlon in dealers:
                dist = haversine_distance(dlat, dlon, lat, lon)
                ranked.append((dist, name, dlat, dlon))
            ranked.sort(key=lambda t: t[0])
            candidates = ranked[:topk]

            best_name = None
            best_duration_s = None
            best_dist_km = None
            for dist, name, dlat, dlon in candidates:
                duration_s = _osrm_route_duration_s(lat, lon, dlat, dlon, routes_cache)
                if duration_s is None:
                    continue
                routes_cache_dirty = True
                if best_duration_s is None or duration_s < best_duration_s:
                    best_duration_s = duration_s
                    best_name = name
                    best_dist_km = dist

            if best_name is None:
                out = dict(lead)
                out["unassigned_reason"] = "no_route_found"
                unassigned_rows.append(out)
                continue

            if max_hours is not None and best_duration_s is not None:
                if best_duration_s > max_hours * 3600:
                    out = dict(lead)
                    out["unassigned_reason"] = "over_max_hours"
                    out["drive_time_hr"] = round(best_duration_s / 3600.0, 2)
                    unassigned_rows.append(out)
                    continue

            out = dict(lead)
            out.setdefault("tag_names", [])
            out["nearest_dealer"] = best_name
            if best_dist_km is not None:
                out["distance_km"] = round(best_dist_km, 1)
            if best_duration_s is not None:
                out["drive_time_hr"] = round(best_duration_s / 3600.0, 2)
            assigned_rows.append(out)
            counts[best_name] = counts.get(best_name, 0) + 1
        else:
            best_name = None
            best_dist = 1e18
            for name, dlat, dlon in dealers:
                dist = haversine_distance(dlat, dlon, lat, lon)
                if dist < best_dist:
                    best_dist = dist
                    best_name = name

            if best_name is None:
                out = dict(lead)
                out["unassigned_reason"] = "no_dealer_found"
                unassigned_rows.append(out)
                continue

            if best_dist <= radius_km:
                out = dict(lead)
                out.setdefault("tag_names", [])
                out["nearest_dealer"] = best_name
                out["distance_km"] = round(best_dist, 1)
                assigned_rows.append(out)
                counts[best_name] = counts.get(best_name, 0) + 1
            else:
                out = dict(lead)
                out["unassigned_reason"] = "over_radius_km"
                out["distance_km"] = round(best_dist, 1)
                unassigned_rows.append(out)

        if i % 250 == 0:
            print(f"Processed {i}/{len(resolved)} leads...")

    if use_driving and routes_cache_dirty:
        _save_routes_cache(routes_cache)

    # 2b) Optionally fetch email + tags from Odoo for the leads we actually kept
    if args.fetch_odoo:
        lead_ids: List[int] = []
        for r in assigned_rows:
            lid = get_lead_id(r)
            if lid is not None:
                lead_ids.append(int(lid))
        lead_ids = sorted(set(lead_ids))
        print(f"Fetching email and tags from Odoo for {len(lead_ids)} leads...")
        if lead_ids:
            email_by_id, tagids_by_id = fetch_lead_fields_from_odoo(lead_ids)
            all_tag_ids = sorted({tid for tids in tagids_by_id.values() for tid in tids})
            print(f"Fetching {len(all_tag_ids)} unique tag names from Odoo...")
            tagname_by_id = fetch_tag_names_from_odoo(all_tag_ids) if all_tag_ids else {}
            for r in assigned_rows:
                lid = get_lead_id(r)
                if lid is None:
                    continue
                r["email_from"] = email_by_id.get(lid, r.get("email_from", "") or "")
                tids = tagids_by_id.get(lid, [])
                r["tag_ids"] = tids
                r["tag_names"] = [tagname_by_id.get(t, str(t)) for t in tids if t in tagname_by_id]
        else:
            print("No lead IDs found in leads_export.json; cannot fetch email/tags from Odoo. Ensure export includes lead ID.")

    # 3) Write Excel
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_assigned = wb.create_sheet("Assigned")
    ws_unassigned = wb.create_sheet("Unassigned")

    ws_summary.append(["Dealer", f"Assigned leads within {radius_km:.1f} km"])
    for dealer_name in sorted(counts.keys()):
        ws_summary.append([dealer_name, counts[dealer_name]])

    # Build one column per tag (based on tags present in assigned rows)
    tag_names = sorted({t for r in assigned_rows for t in (r.get("tag_names") or []) if t})
    if use_driving:
        ws_assigned.append(["nearest_dealer", "distance_km", "drive_time_hr", "email_from", *tag_names, *[c for c in LEAD_COLUMNS if c != "email_from"]])
    else:
        ws_assigned.append(["nearest_dealer", "distance_km", "email_from", *tag_names, *[c for c in LEAD_COLUMNS if c != "email_from"]])
    for row in assigned_rows:
        row_tags = set(row.get("tag_names") or [])
        tag_bits = [1 if t in row_tags else 0 for t in tag_names]
        # email_from gets its own explicit column (and is also present in LEAD_COLUMNS in some exports)
        email = row.get("email_from", "")
        lead_vals = []
        for c in LEAD_COLUMNS:
            if c == "email_from":
                continue
            v = row.get(c, "")
            lead_vals.append(v if not isinstance(v, (dict, list)) else str(v))
        if use_driving:
            ws_assigned.append([row.get("nearest_dealer", ""), row.get("distance_km", ""), row.get("drive_time_hr", ""), email, *tag_bits, *lead_vals])
        else:
            ws_assigned.append([row.get("nearest_dealer", ""), row.get("distance_km", ""), email, *tag_bits, *lead_vals])

    # Unassigned sheet (no dealer assignment)
    ws_unassigned.append(["unassigned_reason", "distance_km", "drive_time_hr", "email_from", *tag_names, *[c for c in LEAD_COLUMNS if c != "email_from"]])
    for row in unassigned_rows:
        row_tags = set(row.get("tag_names") or [])
        tag_bits = [1 if t in row_tags else 0 for t in tag_names]
        email = row.get("email_from", "")
        lead_vals = []
        for c in LEAD_COLUMNS:
            if c == "email_from":
                continue
            v = row.get(c, "")
            lead_vals.append(v if not isinstance(v, (dict, list)) else str(v))
        ws_unassigned.append([row.get("unassigned_reason", ""), row.get("distance_km", ""), row.get("drive_time_hr", ""), email, *tag_bits, *lead_vals])

    ws_summary.freeze_panes = "A2"
    ws_assigned.freeze_panes = "A2"
    ws_unassigned.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
