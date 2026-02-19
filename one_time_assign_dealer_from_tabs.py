#!/usr/bin/env python3
"""One-time tool: assign Dealer property on CRM leads from Excel tabs (excluding first tab)."""

import argparse
import json
import os
import re
import sys
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

GENERIC_TOKENS = {
    "agro", "centre", "center", "home", "farm", "supply", "hardware",
    "coop", "co", "op", "cooperative", "and", "gas", "bar", "bulk",
    "petroleum", "cardlock", "food", "store",
}


def _normalize_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _norm_key(value: Optional[str]) -> str:
    return _normalize_text(value).casefold()


def _norm_phone(value: Optional[str]) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value) if ch.isdigit())


def _compact_alnum(value: Optional[str]) -> str:
    return "".join(ch for ch in _norm_key(value) if ch.isalnum())


def _tokens(value: Optional[str]) -> List[str]:
    base = _norm_key(value).replace("co-op", "coop")
    return [t for t in re.findall(r"[a-z0-9]+", base) if t]


def _info_tokens(value: Optional[str]) -> List[str]:
    return [t for t in _tokens(value) if t not in GENERIC_TOKENS and len(t) >= 3]


def _load_default_credentials(
    url_override: Optional[str] = None,
    db_override: Optional[str] = None,
    username_override: Optional[str] = None,
    password_override: Optional[str] = None,
) -> Tuple[str, str, str, str]:
    url = url_override or os.getenv("ODOO_URL")
    db = db_override or os.getenv("ODOO_DB")
    username = username_override or os.getenv("ODOO_USERNAME")
    password = password_override or os.getenv("ODOO_PASSWORD")

    if url and db and username and password:
        return url, db, username, password

    try:
        import odoo_connector as oc  # noqa: PLC0415

        return (
            url or oc.ODOO_URL,
            db or oc.ODOO_DB,
            username or oc.ODOO_USERNAME,
            password or oc.ODOO_PASSWORD,
        )
    except Exception:
        missing = [k for k, v in {
            "ODOO_URL": url,
            "ODOO_DB": db,
            "ODOO_USERNAME": username,
            "ODOO_PASSWORD": password,
        }.items() if not v]
        raise RuntimeError(
            "Missing Odoo credentials. Set env vars "
            f"{', '.join(missing)} or keep odoo_connector.py available."
        )


class JsonRpcClient:
    def __init__(self, base_url: str, db: str, username: str, password: str):
        self.url = base_url.rstrip("/") + "/jsonrpc"
        self.db = db
        self.username = username
        self.password = password
        self.uid = self._call("common", "login", db, username, password)
        if not self.uid:
            raise RuntimeError("Authentication failed.")

    def _call(self, service: str, method: str, *args):
        payload = {
            "jsonrpc": "2.0",
            "method": "call",
            "params": {"service": service, "method": method, "args": list(args)},
            "id": 1,
        }
        request = urllib.request.Request(
            self.url,
            data=json.dumps(payload).encode(),
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(request, timeout=60) as response:
            data = json.loads(response.read().decode())
        if data.get("error"):
            raise RuntimeError(data["error"])
        return data["result"]

    def execute_kw(self, model: str, method: str, args: list, kwargs: Optional[dict] = None):
        if kwargs is None:
            kwargs = {}
        return self._call(
            "object",
            "execute_kw",
            self.db,
            self.uid,
            self.password,
            model,
            method,
            args,
            kwargs,
        )


@dataclass
class RowItem:
    sheet: str
    row: int
    dealer_snippet: str
    name: str
    email: str
    phone_digits: str
    mobile_digits: str
    city: str


def write_unresolved_report(path: str, rows: List[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Unresolved"
    headers = [
        "sheet",
        "row",
        "reason",
        "lead_match_method",
        "dealer_match_method",
        "lead_id",
        "name",
        "email",
        "phone_digits",
        "mobile_digits",
        "city",
        "dealer_snippet",
    ]
    ws.append(headers)
    for item in rows:
        ws.append([item.get(h, "") for h in headers])
    wb.save(path)


def read_rows_from_tabs(excel_path: str) -> List[RowItem]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    rows: List[RowItem] = []
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        if sheet_idx == 0:
            continue  # ignore first tab
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            dealer = _normalize_text(ws.cell(row=r, column=1).value)
            name = _normalize_text(ws.cell(row=r, column=3).value)
            phone = _norm_phone(ws.cell(row=r, column=4).value)
            mobile = _norm_phone(ws.cell(row=r, column=5).value)
            city = _normalize_text(ws.cell(row=r, column=6).value)
            email = _normalize_text(ws.cell(row=r, column=8).value).lower()
            if not any([dealer, name, phone, mobile, email]):
                continue
            rows.append(
                RowItem(
                    sheet=sheet_name,
                    row=r,
                    dealer_snippet=dealer,
                    name=name,
                    email=email,
                    phone_digits=phone,
                    mobile_digits=mobile,
                    city=city,
                )
            )
    return rows


def load_leads_basic(rpc: JsonRpcClient) -> List[dict]:
    ids = rpc.execute_kw("crm.lead", "search", [[["active", "in", [True, False]]]], {"limit": 0})
    if not ids:
        return []
    leads = rpc.execute_kw(
        "crm.lead",
        "read",
        [ids],
        {"fields": ["id", "name", "email_from", "phone", "mobile", "city"]},
    )
    return leads


def build_indices(leads: List[dict]):
    by_email = defaultdict(list)
    by_phone = defaultdict(list)
    by_name_city = defaultdict(list)
    by_name = defaultdict(list)
    for lead in leads:
        if lead.get("email_from"):
            by_email[_norm_key(lead["email_from"])].append(lead)
        for p in (lead.get("phone"), lead.get("mobile")):
            np = _norm_phone(p)
            if np:
                by_phone[np].append(lead)
        name_key = _norm_key(lead.get("name"))
        city_key = _norm_key(lead.get("city"))
        if name_key:
            by_name[name_key].append(lead)
            if city_key:
                by_name_city[(name_key, city_key)].append(lead)
    return by_email, by_phone, by_name_city, by_name


def match_lead(row: RowItem, indices) -> Tuple[Optional[int], str]:
    by_email, by_phone, by_name_city, by_name = indices
    saw_ambiguous = False

    if row.email:
        candidates = by_email.get(_norm_key(row.email), [])
        if len(candidates) == 1:
            return candidates[0]["id"], "email"
        if len(candidates) > 1:
            saw_ambiguous = True

    phones = [p for p in [row.phone_digits, row.mobile_digits] if p]
    for p in phones:
        candidates = by_phone.get(p, [])
        if len(candidates) == 1:
            return candidates[0]["id"], "phone"
        if len(candidates) > 1:
            saw_ambiguous = True

    nk = _norm_key(row.name)
    ck = _norm_key(row.city)
    if nk and ck:
        candidates = by_name_city.get((nk, ck), [])
        if len(candidates) == 1:
            return candidates[0]["id"], "name_city"
        if len(candidates) > 1:
            saw_ambiguous = True

    if nk:
        candidates = by_name.get(nk, [])
        if len(candidates) == 1:
            return candidates[0]["id"], "name"
        if len(candidates) > 1:
            saw_ambiguous = True

    if saw_ambiguous:
        return None, "ambiguous"
    return None, "no_match"


def load_dealer_option_map(rpc: JsonRpcClient, property_label: str, team_id: Optional[int]) -> Dict[str, str]:
    team_domain = [["id", "=", team_id]] if team_id else []
    teams = rpc.execute_kw(
        "crm.team",
        "search_read",
        [team_domain],
        {"fields": ["id", "name", "lead_properties_definition"], "order": "id asc"},
    )
    for team in teams:
        defs = team.get("lead_properties_definition") or []
        for item in defs:
            if not isinstance(item, dict):
                continue
            if item.get("type") != "selection":
                continue
            if _norm_key(item.get("string")) != _norm_key(property_label):
                continue
            options = item.get("selection") or []
            option_map = {}
            for option in options:
                if isinstance(option, (list, tuple)) and len(option) >= 2:
                    option_map[_normalize_text(option[1])] = str(option[0])
            if option_map:
                return option_map
    raise RuntimeError(f"Could not find '{property_label}' selection property with options in crm.team.")


def pick_dealer_option_key(snippet: str, option_map: Dict[str, str]) -> Tuple[Optional[str], str]:
    s = _norm_key(snippet)
    if not s:
        return None, "empty_snippet"

    exact = [label for label in option_map if _norm_key(label) == s]
    if len(exact) == 1:
        return option_map[exact[0]], "exact"
    if len(exact) > 1:
        return None, "ambiguous_exact"

    s_compact = _compact_alnum(snippet)
    contains = [label for label in option_map if s in _norm_key(label) or (s_compact and s_compact in _compact_alnum(label))]
    if len(contains) == 1:
        return option_map[contains[0]], "contains"
    if len(contains) > 1:
        # resolve using longest label when all contain the same compact snippet
        contains_sorted = sorted(contains, key=lambda x: len(_norm_key(x)))
        return option_map[contains_sorted[-1]], "contains_longest"

    reverse_contains = [label for label in option_map if _norm_key(label) in s or (_compact_alnum(label) and _compact_alnum(label) in s_compact)]
    if len(reverse_contains) == 1:
        return option_map[reverse_contains[0]], "reverse_contains"
    if len(reverse_contains) > 1:
        return None, "ambiguous_reverse_contains"

    # Unique long token in snippet can disambiguate (e.g. "Lang Gas Bar" -> "PRAIRIE SKY - Lang Agro Centre")
    snippet_tokens = [t for t in _info_tokens(snippet) if len(t) >= 4]
    if snippet_tokens:
        token_hits = {}
        for token in snippet_tokens:
            matches = [label for label in option_map if token in _info_tokens(label)]
            token_hits[token] = matches
        unique_token_labels = [matches[0] for matches in token_hits.values() if len(matches) == 1]
        unique_token_labels = list(dict.fromkeys(unique_token_labels))
        if len(unique_token_labels) == 1:
            return option_map[unique_token_labels[0]], "unique_token"
        if len(unique_token_labels) > 1:
            return None, "ambiguous_unique_token"

    # Fuzzy token overlap for labels like "Kindersley Co-op Brock Cardlock"
    s_tokens = set(_info_tokens(snippet))
    if s_tokens:
        scored = []
        for label in option_map:
            l_tokens = set(_info_tokens(label))
            if not l_tokens:
                continue
            inter = len(s_tokens & l_tokens)
            if inter == 0:
                continue
            score = inter / max(1, len(s_tokens))
            if score >= 0.5:
                scored.append((score, inter, len(l_tokens), label))
        if scored:
            scored.sort(reverse=True)
            best = scored[0]
            if len(scored) == 1:
                return option_map[best[3]], "token_fuzzy"
            second = scored[1]
            # require clear win to avoid incorrect assignments
            if best[0] >= second[0] + 0.2 or (best[0] == 1.0 and best[1] > second[1]):
                return option_map[best[3]], "token_fuzzy_clear"
            return None, "ambiguous_fuzzy"

    return None, "no_option_match"


def build_updates(
    rpc: JsonRpcClient,
    matched: Dict[int, str],
    property_label: str,
) -> Tuple[List[Tuple[int, list]], int]:
    lead_ids = sorted(matched.keys())
    if not lead_ids:
        return [], 0

    rows = rpc.execute_kw(
        "crm.lead",
        "read",
        [lead_ids],
        {"fields": ["id", "lead_properties"]},
    )
    updates = []
    already = 0
    for row in rows:
        lead_id = row["id"]
        target_value = matched[lead_id]
        props = row.get("lead_properties") or []
        changed = False
        for item in props:
            if not isinstance(item, dict):
                continue
            if item.get("type") != "selection":
                continue
            if _norm_key(item.get("string")) != _norm_key(property_label):
                continue
            if str(item.get("value") or "") == target_value:
                already += 1
                changed = False
                break
            item["value"] = target_value
            changed = True
            break
        if changed:
            updates.append((lead_id, props))
    return updates, already


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Assign Dealer property values on leads from all tabs except first tab."
    )
    ap.add_argument("--excel", default="Wavcor Summary Leads.xlsx", help="Source Excel workbook")
    ap.add_argument("--property-label", default="Dealer", help="Lead property label")
    ap.add_argument("--team-id", type=int, default=None, help="Optional CRM team id to source options")
    ap.add_argument("--apply", action="store_true", help="Write to Odoo (default is dry-run)")
    ap.add_argument("--limit", type=int, default=0, help="Optional max Excel rows to process")
    ap.add_argument(
        "--unresolved-out",
        default=None,
        help="Optional output .xlsx path for unresolved/skipped rows",
    )
    ap.add_argument("--odoo-url", default=None)
    ap.add_argument("--odoo-db", default=None)
    ap.add_argument("--odoo-username", default=None)
    ap.add_argument("--odoo-password", default=None)
    args = ap.parse_args()

    try:
        url, db, username, password = _load_default_credentials(
            url_override=args.odoo_url,
            db_override=args.odoo_db,
            username_override=args.odoo_username,
            password_override=args.odoo_password,
        )
        rpc = JsonRpcClient(url, db, username, password)

        excel_rows = read_rows_from_tabs(args.excel)
        if args.limit > 0:
            excel_rows = excel_rows[: args.limit]
        print(f"Excel rows read from tabs (excluding first): {len(excel_rows)}")

        leads = load_leads_basic(rpc)
        print(f"Leads loaded from Odoo: {len(leads)}")
        indices = build_indices(leads)
        option_map = load_dealer_option_map(rpc, args.property_label, args.team_id)
        print(f"Dealer options available: {len(option_map)}")

        matched_lead_to_value: Dict[int, str] = {}
        conflicts = 0
        no_lead_match = 0
        no_dealer_match = 0
        ambiguous = 0
        unresolved_rows: List[dict] = []
        sample_no_lead = []
        sample_no_dealer = []
        sample_ambiguous = []
        sample_conflict = []

        for item in excel_rows:
            lead_id, lead_method = match_lead(item, indices)
            if not lead_id:
                if lead_method.startswith("ambiguous"):
                    ambiguous += 1
                    unresolved_rows.append({
                        "sheet": item.sheet,
                        "row": item.row,
                        "reason": "ambiguous_lead_match",
                        "lead_match_method": lead_method,
                        "dealer_match_method": "",
                        "lead_id": "",
                        "name": item.name,
                        "email": item.email,
                        "phone_digits": item.phone_digits,
                        "mobile_digits": item.mobile_digits,
                        "city": item.city,
                        "dealer_snippet": item.dealer_snippet,
                    })
                    if len(sample_ambiguous) < 10:
                        sample_ambiguous.append((item.sheet, item.row, item.name, item.email, item.phone_digits, item.dealer_snippet, lead_method))
                else:
                    no_lead_match += 1
                    unresolved_rows.append({
                        "sheet": item.sheet,
                        "row": item.row,
                        "reason": "no_lead_match",
                        "lead_match_method": lead_method,
                        "dealer_match_method": "",
                        "lead_id": "",
                        "name": item.name,
                        "email": item.email,
                        "phone_digits": item.phone_digits,
                        "mobile_digits": item.mobile_digits,
                        "city": item.city,
                        "dealer_snippet": item.dealer_snippet,
                    })
                    if len(sample_no_lead) < 10:
                        sample_no_lead.append((item.sheet, item.row, item.name, item.email, item.phone_digits, item.dealer_snippet, lead_method))
                continue

            option_key, option_method = pick_dealer_option_key(item.dealer_snippet, option_map)
            if not option_key:
                if option_method.startswith("ambiguous"):
                    ambiguous += 1
                    unresolved_rows.append({
                        "sheet": item.sheet,
                        "row": item.row,
                        "reason": "ambiguous_dealer_match",
                        "lead_match_method": lead_method,
                        "dealer_match_method": option_method,
                        "lead_id": lead_id,
                        "name": item.name,
                        "email": item.email,
                        "phone_digits": item.phone_digits,
                        "mobile_digits": item.mobile_digits,
                        "city": item.city,
                        "dealer_snippet": item.dealer_snippet,
                    })
                    if len(sample_ambiguous) < 10:
                        sample_ambiguous.append((item.sheet, item.row, item.name, item.email, item.phone_digits, item.dealer_snippet, option_method))
                else:
                    no_dealer_match += 1
                    unresolved_rows.append({
                        "sheet": item.sheet,
                        "row": item.row,
                        "reason": "no_dealer_match",
                        "lead_match_method": lead_method,
                        "dealer_match_method": option_method,
                        "lead_id": lead_id,
                        "name": item.name,
                        "email": item.email,
                        "phone_digits": item.phone_digits,
                        "mobile_digits": item.mobile_digits,
                        "city": item.city,
                        "dealer_snippet": item.dealer_snippet,
                    })
                    if len(sample_no_dealer) < 10:
                        sample_no_dealer.append((item.sheet, item.row, item.name, item.email, item.phone_digits, item.dealer_snippet, option_method))
                continue

            existing = matched_lead_to_value.get(lead_id)
            if existing and existing != option_key:
                conflicts += 1
                unresolved_rows.append({
                    "sheet": item.sheet,
                    "row": item.row,
                    "reason": "conflict_same_lead_multiple_dealers",
                    "lead_match_method": lead_method,
                    "dealer_match_method": option_method,
                    "lead_id": lead_id,
                    "name": item.name,
                    "email": item.email,
                    "phone_digits": item.phone_digits,
                    "mobile_digits": item.mobile_digits,
                    "city": item.city,
                    "dealer_snippet": item.dealer_snippet,
                })
                if len(sample_conflict) < 10:
                    sample_conflict.append((item.sheet, item.row, lead_id, item.name, item.dealer_snippet))
                continue
            matched_lead_to_value[lead_id] = option_key

        updates, already_set = build_updates(rpc, matched_lead_to_value, args.property_label)

        print(f"Matched leads with a dealer option: {len(matched_lead_to_value)}")
        print(f"Updates required (dealer value differs): {len(updates)}")
        print(f"Already set correctly: {already_set}")
        print(f"No lead match: {no_lead_match}")
        print(f"No dealer option match: {no_dealer_match}")
        print(f"Ambiguous rows: {ambiguous}")
        print(f"Conflicting rows for same lead: {conflicts}")
        print(f"Total unresolved/skipped rows: {len(unresolved_rows)}")
        if args.unresolved_out:
            write_unresolved_report(args.unresolved_out, unresolved_rows)
            print(f"Unresolved report written: {args.unresolved_out}")
        if sample_no_lead:
            print("Sample no-lead-match rows:")
            for s in sample_no_lead:
                print(f"  {s}")
        if sample_no_dealer:
            print("Sample no-dealer-match rows:")
            for s in sample_no_dealer:
                print(f"  {s}")
        if sample_ambiguous:
            print("Sample ambiguous rows:")
            for s in sample_ambiguous:
                print(f"  {s}")
        if sample_conflict:
            print("Sample conflicting rows:")
            for s in sample_conflict:
                print(f"  {s}")

        if not args.apply:
            print("Dry run only. Re-run with --apply to write Dealer values.")
            return 0

        changed = 0
        for lead_id, props in updates:
            rpc.execute_kw("crm.lead", "write", [[lead_id], {"lead_properties": props}])
            changed += 1
        print(f"Leads updated: {changed}")
        return 0
    except FileNotFoundError as exc:
        print(f"ERROR: Excel file not found: {exc}")
        return 1
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 2


if __name__ == "__main__":
    sys.exit(main())
