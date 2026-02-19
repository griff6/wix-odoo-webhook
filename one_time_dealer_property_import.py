#!/usr/bin/env python3
"""One-time tool: add Dealer property options (lead_properties) from Excel Column A."""

import argparse
import os
import re
import sys
import uuid
import xmlrpc.client
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def _norm_key(value: str) -> str:
    return _normalize_text(value).casefold()


def _load_default_credentials(
    url_override: Optional[str] = None,
    db_override: Optional[str] = None,
    username_override: Optional[str] = None,
    password_override: Optional[str] = None,
) -> Tuple[str, str, str, str]:
    url = url_override or os.getenv("ODOO_URL")
    db = db_override or os.getenv("ODOO_DB")
    username = username_override or os.getenv("ODOO_USERNAME")
    password = password_override or os.getenv("ODOO_PASSWORD")

    if url and db and username and password:
        return url, db, username, password

    try:
        import odoo_connector as oc  # noqa: PLC0415

        return (url or oc.ODOO_URL, db or oc.ODOO_DB, username or oc.ODOO_USERNAME, password or oc.ODOO_PASSWORD)
    except Exception:
        missing = [k for k, v in {
            "ODOO_URL": url,
            "ODOO_DB": db,
            "ODOO_USERNAME": username,
            "ODOO_PASSWORD": password,
        }.items() if not v]
        raise RuntimeError(
            "Missing Odoo credentials. Set env vars "
            f"{', '.join(missing)} or keep odoo_connector.py available."
        )


def connect_odoo(
    url: str,
    db: str,
    username: str,
    password: str,
) -> Tuple[int, xmlrpc.client.ServerProxy]:
    common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", allow_none=True)
    uid = common.authenticate(db, username, password, {})
    if not uid:
        raise RuntimeError("Authentication failed. Check Odoo credentials.")
    models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", allow_none=True)
    return uid, models


def read_unique_values_from_column_a(
    file_path: str,
    sheet_name: Optional[str] = None,
    skip_header: bool = True,
) -> List[str]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    unique: Dict[str, str] = {}
    for row_index, row in enumerate(
        ws.iter_rows(min_col=1, max_col=1, values_only=True),
        start=1,
    ):
        if skip_header and row_index == 1:
            continue
        raw = row[0]
        if raw is None:
            continue
        value = _normalize_text(str(raw))
        if not value:
            continue
        key = _norm_key(value)
        if key not in unique:
            unique[key] = value
    return list(unique.values())


def _extract_option_label(option) -> str:
    if isinstance(option, (list, tuple)) and len(option) >= 2:
        return _normalize_text(str(option[1]))
    if isinstance(option, dict):
        for k in ("label", "name", "string", "value"):
            if option.get(k):
                return _normalize_text(str(option[k]))
    if isinstance(option, str):
        return _normalize_text(option)
    return ""


def _build_option_like(example_option, label: str):
    # Try to preserve existing payload style.
    if isinstance(example_option, (list, tuple)):
        key = uuid.uuid4().hex[:16]
        return [key, label]
    if isinstance(example_option, dict):
        result = dict(example_option)
        key = uuid.uuid4().hex[:16]
        if "name" in result:
            result["name"] = label
        if "label" in result:
            result["label"] = label
        if "string" in result:
            result["string"] = label
        if "value" in result and not result.get("value"):
            result["value"] = key
        return result
    key = uuid.uuid4().hex[:16]
    return [key, label]


def get_target_teams(
    models: xmlrpc.client.ServerProxy,
    db: str,
    uid: int,
    password: str,
    team_id: Optional[int],
    team_name: Optional[str],
) -> List[dict]:
    if team_id:
        domain = [["id", "=", team_id]]
    elif team_name:
        domain = [["name", "ilike", team_name]]
    else:
        domain = []
    return models.execute_kw(
        db,
        uid,
        password,
        "crm.team",
        "search_read",
        [domain],
        {"fields": ["id", "name", "lead_properties_definition"], "order": "id asc"},
    )


def find_property_definition(definitions: list, property_label: str) -> Optional[Tuple[int, dict]]:
    target = property_label.casefold()
    for i, item in enumerate(definitions or []):
        if not isinstance(item, dict):
            continue
        if item.get("type") != "selection":
            continue
        if str(item.get("string") or "").casefold() == target:
            return i, item
    return None


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Add missing Dealer options to CRM Lead Properties from Excel Column A."
    )
    ap.add_argument("--excel", default="Wavcor Summary Leads.xlsx", help="Path to Excel file")
    ap.add_argument("--sheet", default=None, help="Optional sheet name")
    ap.add_argument("--include-header", action="store_true", help="Include row 1 from Column A")
    ap.add_argument("--property-label", default="Dealer", help="Lead property label to update")
    ap.add_argument("--team-id", type=int, default=None, help="Optional target CRM team ID")
    ap.add_argument("--team-name", default=None, help="Optional target CRM team name")
    ap.add_argument("--apply", action="store_true", help="Write changes to Odoo")
    ap.add_argument("--odoo-url", default=None, help="Override Odoo URL")
    ap.add_argument("--odoo-db", default=None, help="Override Odoo database name")
    ap.add_argument("--odoo-username", default=None, help="Override Odoo username")
    ap.add_argument("--odoo-password", default=None, help="Override Odoo password")
    args = ap.parse_args()

    try:
        excel_values = read_unique_values_from_column_a(
            file_path=args.excel,
            sheet_name=args.sheet,
            skip_header=not args.include_header,
        )
        print(f"Read {len(excel_values)} unique non-empty values from Column A.")
        if not excel_values:
            print("No values found; nothing to do.")
            return 0

        url, db, username, password = _load_default_credentials(
            url_override=args.odoo_url,
            db_override=args.odoo_db,
            username_override=args.odoo_username,
            password_override=args.odoo_password,
        )
        uid, models = connect_odoo(url=url, db=db, username=username, password=password)

        teams = get_target_teams(
            models=models,
            db=db,
            uid=uid,
            password=password,
            team_id=args.team_id,
            team_name=args.team_name,
        )
        if not teams:
            raise RuntimeError("No CRM teams found for the provided filter.")

        updated_teams = 0
        for team in teams:
            defs = team.get("lead_properties_definition") or []
            found = find_property_definition(definitions=defs, property_label=args.property_label)
            if not found:
                print(f"Team {team['id']} ({team['name']}): property '{args.property_label}' not found. Skipping.")
                continue

            prop_index, prop_def = found
            selection = prop_def.get("selection") or []
            existing_labels = {
                _norm_key(_extract_option_label(opt))
                for opt in selection
                if _extract_option_label(opt)
            }
            missing = [v for v in excel_values if _norm_key(v) not in existing_labels]
            print(f"Team {team['id']} ({team['name']}): existing={len(selection)} missing={len(missing)}")

            if missing:
                for value in missing:
                    print(f"  - {value}")

            if not args.apply or not missing:
                continue

            example = selection[0] if selection else ["example_key", "Example Label"]
            new_selection = list(selection)
            for label in missing:
                new_selection.append(_build_option_like(example, label))

            new_defs = list(defs)
            new_prop = dict(prop_def)
            new_prop["selection"] = new_selection
            new_defs[prop_index] = new_prop

            models.execute_kw(
                db,
                uid,
                password,
                "crm.team",
                "write",
                [[team["id"]], {"lead_properties_definition": new_defs}],
            )
            updated_teams += 1
            print(f"Team {team['id']} updated with {len(missing)} new options.")

        if not args.apply:
            print("Dry run only. Re-run with --apply to update Odoo.")
        else:
            print(f"Updated teams: {updated_teams}")
        return 0

    except FileNotFoundError as exc:
        print(f"ERROR: Excel file not found: {exc}")
        return 1
    except xmlrpc.client.Fault as exc:
        print(f"Odoo RPC ERROR: {exc.faultString}")
        return 2
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 3


if __name__ == "__main__":
    sys.exit(main())
