#!/usr/bin/env python3
"""One-time tool: create Odoo CRM dealer selection options from Excel column A."""

import argparse
import os
import re
import sys
import xmlrpc.client
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def _norm_key(value: str) -> str:
    return _normalize_text(value).casefold()


def read_unique_values_from_column_a(
    file_path: str,
    sheet_name: Optional[str] = None,
    skip_header: bool = True,
) -> List[str]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    unique: Dict[str, str] = {}
    for row_index, row in enumerate(
        ws.iter_rows(min_col=1, max_col=1, values_only=True),
        start=1,
    ):
        if skip_header and row_index == 1:
            continue
        raw = row[0]
        if raw is None:
            continue
        value = _normalize_text(str(raw))
        if not value:
            continue
        key = _norm_key(value)
        if key not in unique:
            unique[key] = value
    return list(unique.values())


def connect_odoo(
    url: str,
    db: str,
    username: str,
    password: str,
) -> Tuple[int, xmlrpc.client.ServerProxy]:
    common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", allow_none=True)
    uid = common.authenticate(db, username, password, {})
    if not uid:
        raise RuntimeError("Authentication failed. Check Odoo credentials.")
    models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", allow_none=True)
    return uid, models


def find_target_selection_field(
    models: xmlrpc.client.ServerProxy,
    db: str,
    uid: int,
    password: str,
    model_name: str,
    field_name: Optional[str],
    field_label: str,
) -> dict:
    if field_name:
        domain = [["model", "=", model_name], ["ttype", "=", "selection"], ["name", "=", field_name]]
    else:
        domain = [
            ["model", "=", model_name],
            ["ttype", "=", "selection"],
            "|",
            ["field_description", "ilike", field_label],
            ["name", "ilike", field_label],
        ]

    fields = models.execute_kw(
        db,
        uid,
        password,
        "ir.model.fields",
        "search_read",
        [domain],
        {"fields": ["id", "name", "field_description"], "order": "id asc"},
    )

    if not fields:
        raise RuntimeError(
            f"No selection field found on {model_name} "
            f"for field_name={field_name!r} / field_label={field_label!r}."
        )
    if len(fields) > 1 and not field_name:
        choices = ", ".join([f"{f['name']} ({f['field_description']})" for f in fields])
        raise RuntimeError(
            "Multiple matching selection fields found. Re-run with --field-name. "
            f"Matches: {choices}"
        )
    return fields[0]


def fetch_existing_selections(
    models: xmlrpc.client.ServerProxy,
    db: str,
    uid: int,
    password: str,
    field_id: int,
) -> List[dict]:
    return models.execute_kw(
        db,
        uid,
        password,
        "ir.model.fields.selection",
        "search_read",
        [[["field_id", "=", field_id]]],
        {"fields": ["id", "value", "name", "sequence"], "order": "sequence asc, id asc"},
    )


def build_missing_values(excel_values: List[str], existing: List[dict]) -> List[str]:
    existing_keys = set()
    for option in existing:
        name = option.get("name") or ""
        value = option.get("value") or ""
        existing_keys.add(_norm_key(name))
        existing_keys.add(_norm_key(value))
    return [v for v in excel_values if _norm_key(v) not in existing_keys]


def create_selection_options(
    models: xmlrpc.client.ServerProxy,
    db: str,
    uid: int,
    password: str,
    field_id: int,
    values: List[str],
    starting_sequence: int,
) -> int:
    created = 0
    sequence = starting_sequence
    for label in values:
        sequence += 1
        payload = {
            "field_id": field_id,
            "value": label,
            "name": label,
            "sequence": sequence,
        }
        models.execute_kw(
            db,
            uid,
            password,
            "ir.model.fields.selection",
            "create",
            [payload],
        )
        created += 1
    return created


def _load_default_credentials(
    url_override: Optional[str] = None,
    db_override: Optional[str] = None,
    username_override: Optional[str] = None,
    password_override: Optional[str] = None,
) -> Tuple[str, str, str, str]:
    url = url_override or os.getenv("ODOO_URL")
    db = db_override or os.getenv("ODOO_DB")
    username = username_override or os.getenv("ODOO_USERNAME")
    password = password_override or os.getenv("ODOO_PASSWORD")

    if url and db and username and password:
        return url, db, username, password

    try:
        import odoo_connector as oc  # noqa: PLC0415

        return (url or oc.ODOO_URL, db or oc.ODOO_DB, username or oc.ODOO_USERNAME, password or oc.ODOO_PASSWORD)
    except Exception:
        missing = [k for k, v in {
            "ODOO_URL": url,
            "ODOO_DB": db,
            "ODOO_USERNAME": username,
            "ODOO_PASSWORD": password,
        }.items() if not v]
        raise RuntimeError(
            "Missing Odoo credentials. Set env vars "
            f"{', '.join(missing)} or keep odoo_connector.py available."
        )


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Create missing Odoo CRM dealer selection options from Excel column A."
    )
    ap.add_argument(
        "--excel",
        default="Wavcor Summary Leads.xlsx",
        help="Path to source Excel file (default: %(default)s)",
    )
    ap.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name (default: active sheet)",
    )
    ap.add_argument(
        "--include-header",
        action="store_true",
        help="Include row 1 from Column A (default skips header row)",
    )
    ap.add_argument(
        "--model",
        default="crm.lead",
        help="Target Odoo model (default: %(default)s)",
    )
    ap.add_argument(
        "--field-name",
        default=None,
        help="Exact technical field name, e.g. x_studio_dealer",
    )
    ap.add_argument(
        "--field-label",
        default="dealer",
        help="Field label text to find when --field-name is not provided",
    )
    ap.add_argument(
        "--apply",
        action="store_true",
        help="Actually create missing options in Odoo. Without this flag, runs dry-run only.",
    )
    ap.add_argument("--odoo-url", default=None, help="Override Odoo URL, e.g. https://your-db.odoo.com")
    ap.add_argument("--odoo-db", default=None, help="Override Odoo database name")
    ap.add_argument("--odoo-username", default=None, help="Override Odoo username/login")
    ap.add_argument("--odoo-password", default=None, help="Override Odoo password")
    args = ap.parse_args()

    try:
        excel_values = read_unique_values_from_column_a(
            file_path=args.excel,
            sheet_name=args.sheet,
            skip_header=not args.include_header,
        )
        print(f"Read {len(excel_values)} unique non-empty values from Column A.")
        if not excel_values:
            print("No values found; nothing to do.")
            return 0

        url, db, username, password = _load_default_credentials(
            url_override=args.odoo_url,
            db_override=args.odoo_db,
            username_override=args.odoo_username,
            password_override=args.odoo_password,
        )
        uid, models = connect_odoo(url=url, db=db, username=username, password=password)

        field = find_target_selection_field(
            models=models,
            db=db,
            uid=uid,
            password=password,
            model_name=args.model,
            field_name=args.field_name,
            field_label=args.field_label,
        )
        print(
            "Target field: "
            f"{field['name']} ({field['field_description']}) [id={field['id']}] on {args.model}"
        )

        existing = fetch_existing_selections(
            models=models,
            db=db,
            uid=uid,
            password=password,
            field_id=field["id"],
        )
        missing = build_missing_values(excel_values, existing)
        print(f"Existing options: {len(existing)}")
        print(f"Missing options to create: {len(missing)}")

        if missing:
            for value in missing:
                print(f"  - {value}")

        if not args.apply:
            print("Dry run only. Re-run with --apply to create options.")
            return 0

        if not missing:
            print("Nothing to create.")
            return 0

        max_sequence = max([int(opt.get("sequence") or 0) for opt in existing], default=0)
        created_count = create_selection_options(
            models=models,
            db=db,
            uid=uid,
            password=password,
            field_id=field["id"],
            values=missing,
            starting_sequence=max_sequence,
        )
        print(f"Created {created_count} new selection options.")
        return 0
    except FileNotFoundError as exc:
        print(f"ERROR: Excel file not found: {exc}")
        return 1
    except xmlrpc.client.Fault as exc:
        print(f"Odoo RPC ERROR: {exc.faultString}")
        return 2
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 3


if __name__ == "__main__":
    sys.exit(main())
